import csv
import re
import sys
import math
import prettytable
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import openpyxl.styles.numbers
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime


class Salary:
    """Класс для представления зарплаты.

    Attributes:
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница  вилки оклада
        salary_currency (str): Валюята оклада
        salary_gross (str): Указан ли оклад до вычета налогов
        average (int): Средняя зарплата
    """
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Инициализирует объект Salary, выполняет конвертацию для целоисленных полей

        Args:
            salary_from (str or int or float):Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница  вилки оклада
            salary_gross (str or int or float): Валюята оклада
            salary_currency(str): Указан ли оклад до вычета налогов
        """
        self.salary_from = float(salary_from) * self.currency_to_rub[salary_currency]
        self.salary_to = float(salary_to) * self.currency_to_rub[salary_currency]
        self.salary_currency = salary_currency
        self.salary_gross = salary_gross
        self.average = int((self.salary_from + self.salary_to) / 2)


class Vacancy:
    """Класс для представления ваканcии.

    Attributes:
        name(str): Название вакансии
        salary(Salary): Информация о зарплатае вакансии
        area_name(str): Место работы
        published_at(str): Где опубликовано
        description(str): Описание вакансии
        key_skills(list<str>): Необходимые навыки для вакансии
        experience_id(str): Необходимый опыт работы
        premium(str): Премиум
        employer_name(str): Название компании
    """
    def __init__(self, vacancy_information):
        """Инициализирует объект Vacancy

        Args:
             vacancy_information(dict<str:str>): Словарь, в котором записаны значения вакансии
        """
        self.name = vacancy_information['name']
        self.salary = Salary(vacancy_information['salary_from'], vacancy_information['salary_to'],
                             vacancy_information['salary_gross'], vacancy_information['salary_currency'])
        self.area_name = vacancy_information['area_name']
        self.published_at = vacancy_information['published_at']
        self.description = vacancy_information['description']
        self.key_skills = vacancy_information['key_skills'].split(';')
        self.experience_id = vacancy_information['experience_id']
        self.premium = vacancy_information['premium']
        self.employer_name = vacancy_information['employer_name']


class StatisticalDataProcessor:
    """Класс для представления статистики по вакансиям.

    Attributes:
        dataset(list<Vacancy>): Список вакансий
        name_of_profession(str): Название вакансии, по которой будет полученя статистика
        average_salary(list<str>): Динамика уровня зарплат по годам
        average_salary_profession(list<str>): Динамика уровня зарплат по годам для выбранной профессии
        number_of_vacancies(list<str>): Динамика количества вакансий по годам
        number_of_vacancies_profession(list<str>): Динамика количества вакансий по годам для выбранной профессии
        salary_level(list<str>): Уровень зарплат по городам (в порядке убывания)
        vacancy_rate(list<str>):  Доля вакансий по городам (в порядке убывания)
        """
    def __init__(self, dataset):
        """Инициализирует объект StatisticalDataProcessor

        Args:
            dataset(list<Vacancy>): Список вакансий
        """
        self.dataset = dataset
        self.name_of_profession = dataset.name_of_profession

        self.average_salary = self.get_salary_statistics()
        self.average_salary_profession = self.get_salary_statistics(self.name_of_profession)
        self.number_of_vacancies = self.get_number_of_vacancies_statistics()
        self.number_of_vacancies_profession = self.get_number_of_vacancies_statistics(self.name_of_profession)
        self.salary_level = self.get_salary_level_city_statistics()
        self.vacancy_rate = self.get_number_of_vacancies_city_statistics()

    def print_statistic(self):
        """Выводит свю имеющиеся статистику"""
        print('Динамика уровня зарплат по годам: {' + ', '.join(self.average_salary) + '}')
        print('Динамика количества вакансий по годам: {' + ', '.join(self.number_of_vacancies) + '}')
        print('Динамика уровня зарплат по годам для выбранной профессии: {' + ', '.join(
            self.average_salary_profession) + '}')
        print('Динамика количества вакансий по годам для выбранной профессии: {' + ', '.join(
            self.number_of_vacancies_profession) + '}')
        print('Уровень зарплат по городам (в порядке убывания): {' + ', '.join(self.salary_level) + '}')
        print('Доля вакансий по городам (в порядке убывания): {' + ', '.join(self.vacancy_rate) + '}')

    def get_final_year_statistics(self):
        """Конвертирует статистку по зарплате и количеству вакансий в словари
        Returns:
            list<dict<int,int>>: Список статистик по зарплате и количеству вакансий
        """
        return [self.convert_year_statistic_to_dictionary(self.average_salary),
                self.convert_year_statistic_to_dictionary(self.average_salary_profession),
                self.convert_year_statistic_to_dictionary(self.number_of_vacancies),
                self.convert_year_statistic_to_dictionary(self.number_of_vacancies_profession)]

    def get_final_city_statistics(self):
        """Конвертирует статистику по городам в словари
        Returns:
            list<dict<int,float>>: Список статистике по городам
        """
        return [self.convert_city_statistic_to_dictionary(self.salary_level),
                self.convert_city_statistic_to_dictionary(self.vacancy_rate)]

    def get_salary_statistics(self, name_of_profession=""):
        """ Возвращает динамику уровня зарплат по годам, если не передан аргумент name_of_profession.
            Возвращает динамику уровня зарплат по годам для профессии, если передан аргумент name_of_profession.
         Args:
             name_of_profession(str): Название вакансии, по которйо нужно получить статистику
         Returns:
            list<str>: Статистика по уровню зарплат по годам
        """
        salary_statistics = []
        current_year = self.dataset.vacancies_objects[0].published_at.year
        average_salary = 0
        number_of_vacancies = 0
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.published_at.year != current_year and (
                    name_of_profession == "" or name_of_profession in vacancy.name):
                if number_of_vacancies == 0:
                    salary_statistics.append(f'{current_year}: {0}')
                else:
                    salary_statistics.append(f'{current_year}: {math.floor(average_salary / number_of_vacancies)}')
                current_year = vacancy.published_at.year
                average_salary = vacancy.salary.average
                number_of_vacancies = 1
            elif name_of_profession == "" or name_of_profession in vacancy.name:
                average_salary += vacancy.salary.average
                number_of_vacancies += 1
        if number_of_vacancies == 0:
            salary_statistics.append(f'{current_year}: {0}')
        else:
            salary_statistics.append(f'{current_year}: {math.floor(average_salary / number_of_vacancies)}')
        return salary_statistics

    def get_number_of_vacancies_statistics(self, name_of_profession=""):
        """ Возвращает динамику количества вакансий по годам, если не передан аргумент name_of_profession.
            Возвращает динамику уровня зарплат по годам для профессии, если передан аргумент name_of_profession.
        Args:
            name_of_profession(str): Название вакансии, по которйо нужно получить статистику
        Returns:
            list<str>: Статистика по количеству вакансий
        """
        number_of_vacancies_statistics = []
        current_year = self.dataset.vacancies_objects[0].published_at.year
        number_of_vacancies = 0
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.published_at.year != current_year and (
                    name_of_profession == "" or name_of_profession in vacancy.name):
                number_of_vacancies_statistics.append(f'{current_year}: {number_of_vacancies}')
                current_year = vacancy.published_at.year
                number_of_vacancies = 1
            elif name_of_profession == "" or name_of_profession in vacancy.name:
                number_of_vacancies += 1
        number_of_vacancies_statistics.append(f'{current_year}: {number_of_vacancies}')
        return number_of_vacancies_statistics

    def get_salary_level_city_statistics(self):
        """ Возвращает уровень зарплат по городам (в порядке убывания)
        Returns:
            list<str>: Статистика по уровню зарплат по городам
        """
        number_of_vacancies = len(self.dataset.vacancies_objects)
        salary_at_cities = {}
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.area_name not in salary_at_cities:
                salary_at_cities[vacancy.area_name] = (1, vacancy.salary.average)
            else:
                salary_at_cities[vacancy.area_name] = (salary_at_cities[vacancy.area_name][0] + 1,
                                                       salary_at_cities[vacancy.area_name][1] + vacancy.salary.average)
        salary_at_cities = dict(
            sorted(salary_at_cities.items(), key=lambda item: item[1][1] / item[1][0], reverse=True))
        average_salary_at_cities = []
        for city_name, value in salary_at_cities.items():
            if value[0] >= math.floor(number_of_vacancies / 100):
                average_salary_at_cities.append(f"'{city_name}': {math.floor(value[1] / value[0])}")
        return average_salary_at_cities[:10]

    def get_number_of_vacancies_city_statistics(self):
        """ Возвращает долю вакансий по городам (в порядке убывания)
        Returns:
             list<str>: Доля вакансий по городам (в порядке убывания)
        """
        number_of_all_vacancies = len(self.dataset.vacancies_objects)
        number_of_vacancies_at_cities = {}
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.area_name not in number_of_vacancies_at_cities:
                number_of_vacancies_at_cities[vacancy.area_name] = 1
            else:
                number_of_vacancies_at_cities[vacancy.area_name] += 1
        salary_at_cities = dict(sorted(number_of_vacancies_at_cities.items(), key=lambda item: item[1], reverse=True))
        result = []
        for city_name, number_of_vacancies in salary_at_cities.items():
            if number_of_vacancies >= math.floor(number_of_all_vacancies / 100):
                result.append(f"'{city_name}': {round(number_of_vacancies / number_of_all_vacancies, 4)}")
        return result[:10]

    def convert_year_statistic_to_dictionary(self, statistic):
        """ Конвертирует статистку, связянную с зарплатой или с долей вакансий, в словарь
        Args:
            statistic(list<str>):
        Returns:
            dict<int, int>
        """
        return {int(i.split(': ')[0]): int(i.split(': ')[1]) for i in statistic}

    def convert_city_statistic_to_dictionary(self, statistic):
        """ Конвертирует статистку, связянную с городами, в словарь
        Args:
            statistic(list<str>):
        Returns:
            dict<string, float>
        """
        return {i.split(': ')[0][1:-1]: float(i.split(': ')[1]) for i in statistic}


class InputConect:
    """Класс для печатанья таблицы по вакансиям.
        Attributes:
            russian_versions(dict<str, str>): Перевод названия столбцов 
            decoding(dict<str, str>): Перевод значений валюы и опыта с английского на русский
            work_experience_convertor_weight(dict<str, int>): Вес опыта
            formatter_dic(dict<str, str or datetime or list>): Преобразует строку к правильному виду
            filters(dict<str, lambda>): Как должны фильтроваться те или иные поля
            sorting_rules(dict<str, lambda>): Как должны сортироваться те или иные поля
            fields(list<str>): Требуемые столбцы
            vacancy_numbers(int): количество вакансий
        """
    def __init__(self, dataset):
        """Инициализирует класс InputConnect
        Args:
            dataset(Dataset): Информация о вакансиях
        """
        self.russian_versions = {'name': 'Название',
                                 'description': 'Описание',
                                 'key_skills': 'Навыки',
                                 'experience_id': 'Опыт работы',
                                 'premium': 'Премиум-вакансия',
                                 'employer_name': 'Компания',
                                 'salary_from': 'Нижняя граница вилки оклада',
                                 'salary_to': 'Верхняя граница вилки оклада',
                                 'salary_gross': 'Оклад указан до вычета налогов',
                                 'salary_currency': 'Идентификатор валюты оклада',
                                 'area_name': 'Название региона',
                                 'published_at': 'Дата и время публикации вакансии'
                                 }
        self.decoding = {"noExperience": "Нет опыта",
                         "between1And3": "От 1 года до 3 лет",
                         "between3And6": "От 3 до 6 лет",
                         "moreThan6": "Более 6 лет",
                         "AZN": "Манаты",
                         "BYR": "Белорусские рубли",
                         "EUR": "Евро",
                         "GEL": "Грузинский лари",
                         "KGS": "Киргизский сом",
                         "KZT": "Тенге",
                         "RUR": "Рубли",
                         "UAH": "Гривны",
                         "USD": "Доллары",
                         "UZS": "Узбекский сум",
                         'True': 'Да',
                         'False': 'Нет'
                         }
        self.work_experience_convertor_weight = {
            "noExperience": 0,
            "between1And3": 1,
            "between3And6": 2,
            "moreThan6": 3
        }
        self.formatter_dic = {
            'Название': lambda vacancy: vacancy.name,
            'Описание': lambda vacancy: vacancy.description,
            'Навыки': lambda vacancy: ";".join(vacancy.key_skills),
            'Опыт работы': lambda vacancy: self.decoding[vacancy.experience_id],
            'Премиум-вакансия': lambda vacancy: self.decoding[vacancy.premium],
            'Компания': lambda vacancy: vacancy.employer_name,
            'Оклад': lambda vacancy: self.format_salary_information(vacancy),
            'Название региона': lambda vacancy: vacancy.area_name,
            'Дата публикации вакансии': lambda vacancy: datetime.strptime(vacancy.published_at,
                                                                          '%Y-%m-%dT%H:%M:%S%z').strftime(
                '%d.%m.%Y'),
        }
        self.filters = {
            'Компания': lambda employer_name, vacancy: employer_name == vacancy.employer_name,
            'Оклад': lambda salary, vacancy: int(float(vacancy.salary.salary_from)) <= int(float(salary)) <= int(
                float(vacancy.salary.salary_to)),
            'Дата публикации вакансии': lambda date, vacancy: date == datetime.strptime(vacancy.published_at,
                                                                                        '%Y-%m-%dT%H:%M:%S%z').strftime(
                '%d.%m.%Y'),
            'Навыки': lambda skills, vacancy: set(skills.split(', ')).issubset(vacancy.key_skills),
            'Опыт работы': lambda experience_id, vacancy: experience_id == self.decoding[vacancy.experience_id],
            'Премиум-вакансия': lambda premium, vacancy: premium == self.decoding[vacancy.premium],
            'Идентификатор валюты оклада': lambda salary_currency, vacancy: salary_currency == self.decoding[
                vacancy.salary_currency],
            'Название': lambda name, vacancy: name == vacancy.name,
            'Название региона': lambda area_name, vacancy: area_name == vacancy.area_name,
        }
        self.sorting_rules = {
            'Навыки': lambda vacancy: len(vacancy.key_skills),
            'Оклад': lambda vacancy: ((float(vacancy.salary.salary_from) + float(vacancy.salary.salary_to)) / 2) *
                                     self.currency_to_rub[
                                         vacancy.salary_currency],
            'Дата публикации вакансии': lambda vacancy: datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z'),
            'Компания': lambda vacancy: vacancy.employer_name,
            'Опыт работы': lambda vacancy: self.work_experience_convertor_weight[vacancy.experience_id],
            'Премиум-вакансия': lambda vacancy: self.decoding[vacancy.premium],
            'Название региона': lambda vacancy: vacancy.area_name,
            'Название': lambda vacancy: vacancy.name,
            'Идентификатор валюты оклада': lambda vacancy: self.decoding(vacancy.salary_currency)
        }

        self.fields = dataset.fields
        self.vacancy_numbers = dataset.vacancy_numbers
        data_vacancies = self.get_filtered_data_vacancies(dataset.vacancies_objects, dataset.filtering_parameter)
        self.sort_data_vacancies(data_vacancies, dataset.sorting_parameter, dataset.is_reverse_sorting)
        self.table = self.create_table(data_vacancies)

    def get_filtered_data_vacancies(self, dataset, filtering_parameter):
        """Филтурет вакансии по переданному параметру
        Args:
            dataset(list<Vacancy>): Список вакансий
            filtering_parameter(str): Параметр фильтрации
        Resturns:
            list<Vacancy>
        """
        if filtering_parameter != '':
            parameters = filtering_parameter.split(': ')
            filter_name = parameters[0]
            filter_value = parameters[1]
            filtered_data_vacancies = [row for row in dataset if self.filters[filter_name](filter_value, row)]
            if len(filtered_data_vacancies) == 0:
                print('Ничего не найдено')
                sys.exit()
            return filtered_data_vacancies
        return dataset

    def sort_data_vacancies(self, data_vacancies, sorting_parameter, is_reverse_sorting):
        """Сортирует вакансии по переданному параметру
        Args:
            data_vacancies(list<Vacancy>): Список вакансий
            sorting_parameter(str): Параметр сортировки
            is_reverse_sorting(str): Порядок сортировки(Да - прямой, Нет - обратный)
        Resturns:
            list<Vacancy>
        """
        if sorting_parameter != '':
            is_reverse = False if is_reverse_sorting == 'Нет' or is_reverse_sorting == '' else True
            data_vacancies.sort(key=self.sorting_rules[sorting_parameter], reverse=is_reverse)

    def formatter(self, vacancy):
        """Филтурет вакансию к стандартному виду
        Args:
            dataset(Vacancy): вакансия
        Resturns:
                Vacancy
                """
        formatted_vacancy = {}
        for key, value in self.formatter_dic.items():
            formatted_vacancy[key] = value(vacancy)
        return formatted_vacancy

    def format_vacancy_for_table(self, vacancy):
        """Форматирует вакансию к специальному виду для таблицы
        Args:
            vacancy(dict<str, str>): Вакансия
        Resturns:
            dict<str,str>: Форматированная вакансия
        """
        if len(vacancy['Навыки']) > 100:
            vacancy['Навыки'] = vacancy['Навыки'][:100] + '...'
        vacancy['Навыки'] = '\n'.join(vacancy['Навыки'].split(';'))
        for key, value in vacancy.items():
            if key == 'Навыки':
                continue
            if len(value) > 100:
                vacancy[key] = value[:100] + '...'

    def create_table(self, data_vacancies):
        """Создаёт таблицу
        Args:
            data_vacancies(dict<str,str>)
        Returns:
            PrettyTable: Таблица
        """
        if len(data_vacancies) == 0:
            print('Нет данных')
            sys.exit()
        table = PrettyTable()
        columns = ['№'] + list(self.formatter(data_vacancies[0]).keys())
        table.field_names = columns
        for index, vacancy in enumerate(data_vacancies):
            formatted_vacancy = self.formatter(vacancy)
            self.format_vacancy_for_table(formatted_vacancy)
            table.add_row([str(index + 1)] + list(formatted_vacancy.values()))
        table.max_width = 20
        table.align = 'l'
        table.hrules = prettytable.ALL
        return table

    def print_table(self):
        """Печатает таблицу"""
        vacancy_numbers = self.vacancy_numbers.split(' ') if self.vacancy_numbers != '' else []
        fields = ['№'] + self.fields.split(', ') if self.fields != '' else self.table.field_names
        start = int(vacancy_numbers[0]) - 1 if len(vacancy_numbers) >= 1 else 0
        end = int(vacancy_numbers[1]) - 1 if len(vacancy_numbers) == 2 else len(self.table.rows)
        print(self.table.get_string(start=start, end=end, fields=fields))

    def format_salary_information(self, vacancy):
        """Приводит информацию  о зарплате к стандатному виду для таблицы
        Args:
            vacancy(Vacancy): Ваканси
        Resturns:
            str: Отформатированная информация о зарплате
        """
        def format_salary(salary):
            salary = int(float(salary))
            salary = ' '.join(f'{salary:,}'.split(','))
            return salary

        is_salary_gross = 'Без вычета налогов' if vacancy.salary.salary_gross == 'True' else 'С вычетом налогов'
        return f'{format_salary(vacancy.salary.salary_from)} - {format_salary(vacancy.salary.salary_to)} ({self.decoding[vacancy.salary.salary_currency]}) ({is_salary_gross})'


class DataSet:
    """Класс для представления вакансий.

    Attributes:
        file_name(str): Название файла, откуда будут взяты данные для вакансий
        name_of_profession(str): Название професии, по которой можно будет получить список вакансий
        vacancies_objects(list<Vacancy>): Список вакансий
        filtering_parameter(str):параметр фильтрации
        sorting_parameter(str):параметр сортировки
        is_reverse_sorting(str):порядок сортировки
        vacancy_numbers(str): диапазон вывода
        fields(str):требуемые столбцы
    """
    def __init__(self):
        """Инициализирует объект Dataset"""
        self.file_name = input('Введите название файла: ')
        self.name_of_profession = input('Введите название профессии: ')
        self.vacancies_objects = self.__read_data_vacancies()

    def __read_data_vacancies(self):
        """Преобразуе данные из файла в list<Vacancy>
        Returns:
            list<Vacancy>: Список вакансий
        """
        list_naming, list_data = DataSet.csv_reader(self.file_name)
        data_vacancies = DataSet.csv_filer(list_data, list_naming)
        return [Vacancy(vacancy_information) for vacancy_information in data_vacancies]

    def print_table(self):
        """Выводит таблицу по переданным пользователем параметрам
        """
        self.filtering_parameter = input('Введите параметр фильтрации: ')
        self.sorting_parameter = input('Введите параметр сортировки: ')
        self.is_reverse_sorting = input('Обратный порядок сортировки (Да / Нет): ')
        self.vacancy_numbers = input('Введите диапазон вывода: ')
        self.fields = input('Введите требуемые столбцы: ')
        DataSet.check_entered_data(self.filtering_parameter, self.sorting_parameter, self.is_reverse_sorting)

        input_conect = InputConect(dataset)
        input_conect.print_table()

    @staticmethod
    def csv_reader(file_name):
        """Возращает из прочитанного файла список столбоцов и список вакансий
        Args:
            file_name(str): навзание файла
        Returns:
            list<str>
            list<list<str>>
        """
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        list_data = [x for x in reader_csv]
        if len(list_data) == 0:
            print("Пустой файл")
            sys.exit()
        return list_data[0], list_data[1:]

    @staticmethod
    def csv_filer(reader, list_naming):
        """Проверяет столбцы и списки вакансии на ошибки, и преобразует список вакансии в словарь
        Args:
            reader(list<list<str>>): навзание файла
            list_naming(list<str>): Название столбцов
        Returns:
            list<dict<str, str>>: Список вакансий в виде словарей
        """
        rows = [x for x in reader if len(x) == len(list_naming) and '' not in x]
        dictionary_rows = [DataSet.convert_row_to_dictionary(list_naming, row) for row in rows]
        return dictionary_rows

    @staticmethod
    def convert_row_to_dictionary(columns_name, row):
        """Конвертирует список в словарь
        Args:
            columns_name(list<str>): Название столбцов
            row(list<str>): Строка(вакансия)
        Returns:
            dict<str, str>: Вакансия в виде словаря
        """
        dict = {}
        for i in range(len(columns_name)):
            dict[columns_name[i]] = DataSet.convert_cell_to_standard(row[i])
        return dict

    @staticmethod
    def convert_cell_to_standard(cell):
        """Приводит ячейку к стандратному виду
        Args:
            cell(str): ячейка
        Returns:
            cell(str or list<str>): Отформотировання ячейка
        """
        words = re.sub(r'<[^>]*>', '', cell).split('\n')
        for i in range(len(words)):
            words[i] = ' '.join(words[i].split())
        return ';'.join(map(str, words))

    @staticmethod
    def check_entered_data(filtering_parameter, sorting_parameter, is_reverse_sorting):
        """Проверяет введённые параметры вывода таблицы(фильтр, сортировка, порядок) на корректность ввода
         Args:
            filtering_parameter(str): Фильтр
            sorting_parameter(str): Параметр сортировки
            is_reverse_sorting(str): В каком порядке выводится таблица
        """
        currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }
        work_experience_convertor_weight = {
            "noExperience": 0,
            "between1And3": 1,
            "between3And6": 2,
            "moreThan6": 3
        }
        decoding = {"noExperience": "Нет опыта",
                    "between1And3": "От 1 года до 3 лет",
                    "between3And6": "От 3 до 6 лет",
                    "moreThan6": "Более 6 лет",
                    "AZN": "Манаты",
                    "BYR": "Белорусские рубли",
                    "EUR": "Евро",
                    "GEL": "Грузинский лари",
                    "KGS": "Киргизский сом",
                    "KZT": "Тенге",
                    "RUR": "Рубли",
                    "UAH": "Гривны",
                    "USD": "Доллары",
                    "UZS": "Узбекский сум",
                    'True': 'Да',
                    'False': 'Нет'
                    }
        sorting_rules = {
            'Навыки': lambda vacancy: len(vacancy.key_skills),
            'Оклад': lambda vacancy: ((float(vacancy.salary.salary_from) + float(vacancy.salary.salary_to)) / 2) *
                                     currency_to_rub[
                                         vacancy.salary_currency],
            'Дата публикации вакансии': lambda vacancy: datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z'),
            'Компания': lambda vacancy: vacancy.employer_name,
            'Опыт работы': lambda vacancy: work_experience_convertor_weight[vacancy.experience_id],
            'Премиум-вакансия': lambda vacancy: decoding[vacancy.premium],
            'Название региона': lambda vacancy: vacancy.area_name,
            'Название': lambda vacancy: vacancy.name,
            'Идентификатор валюты оклада': lambda vacancy: decoding(vacancy.salary_currency)
        }
        filters = {
            'Компания': lambda employer_name, vacancy: employer_name == vacancy.employer_name,
            'Оклад': lambda salary, vacancy: int(float(vacancy.salary.salary_from)) <= int(float(salary)) <= int(
                float(vacancy.salary.salary_to)),
            'Дата публикации вакансии': lambda date, vacancy: date == datetime.strptime(vacancy.published_at,
                                                                                        '%Y-%m-%dT%H:%M:%S%z').strftime(
                '%d.%m.%Y'),
            'Навыки': lambda skills, vacancy: set(skills.split(', ')).issubset(vacancy.key_skills),
            'Опыт работы': lambda experience_id, vacancy: experience_id == decoding[vacancy.experience_id],
            'Премиум-вакансия': lambda premium, vacancy: premium == decoding[vacancy.premium],
            'Идентификатор валюты оклада': lambda salary_currency, vacancy: salary_currency == decoding[
                vacancy.salary_currency],
            'Название': lambda name, vacancy: name == vacancy.name,
            'Название региона': lambda area_name, vacancy: area_name == vacancy.area_name,
        }
        if filtering_parameter != '' and ':' not in filtering_parameter:
            print('Формат ввода некорректен')
            sys.exit()
        parameters = filtering_parameter.split(': ')
        if filtering_parameter != '' and parameters[0] not in filters.keys():
            print('Параметр поиска некорректен')
            sys.exit()
        if sorting_parameter != '' and is_reverse_sorting not in ('Да', 'Нет', ''):
            print('Порядок сортировки задан некорректно')
            sys.exit()
        if sorting_parameter != '' and sorting_parameter not in sorting_rules.keys():
            print('Параметр сортировки некорректен')
            sys.exit()


class Report:
    """Класс для визуального представления информации о вакансиях.

        Attributes:
            wb(Workbook)
    """
    def __init__(self):
        """Инициализирует объекты Report"""
        self.wb = Workbook()


    def generate_image(self, year_statistics, city_statistics, name_of_profession):
        """Создаёт изображение с диаграммами по вакансиям
        Args:
            year_statistics(list<dict<int, int>>): Статистики по зарплатам и доли вакансий
            city_statistics(list<dict<int, float>>): Статистики по городам
            name_of_profession(str): Название профессии
        """
        fig = plt.figure()
        self.filter_statistics(year_statistics, 2007, 2022)
        self.draw_year_salary_graph(fig, year_statistics[0], year_statistics[1], name_of_profession)
        self.draw_year_vacancy_graph(fig, year_statistics[2], year_statistics[3], name_of_profession)
        self.draw_city_salary_graph(fig, city_statistics[0])
        self.draw_vacancy_rate(fig, city_statistics[1])

        plt.tight_layout()
        plt.savefig("graph.png", dpi=200)

    def draw_year_salary_graph(self, fig, salary_by_years, salary_by_year_profession, name_of_profession):
        """Создаёт диаграмму по уровню зарплат по годам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            salary_by_years(dict<int, int>): Статистика уровня зарплат по годам
            salary_by_year_profession(dict<int, int>):  Статистика уровня зарплат профессии по годам
            name_of_profession(str): Название профессии
        """
        width = 0.4
        x_nums = np.arange(len(salary_by_years.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, salary_by_years.values(), width, label="средняя з/п")
        ax.bar(x_list2, salary_by_year_profession.values(), width, label=f'з/п {name_of_profession.lower()}')

        ax.tick_params(axis="both", labelsize=8)
        ax.set_xticks(x_nums, salary_by_years.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

    def draw_year_vacancy_graph(self, fig, vacancy_by_years, vacancy_by_year_profession, name_of_profession):
        """Создаёт диаграмму по количеству вакансий по годам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            vacancy_by_years(dict<int, int>): Статистика количества вакансий по годам
            vacancy_by_year_profession(dict<int, int>):  Статистика количества вакансий профессии по годам
            name_of_profession(str): Название профессии
        """
        width = 0.4
        x_nums = np.arange(len(vacancy_by_years.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(222)
        ax.set_title("Количество ваканский по годам")
        ax.bar(x_list1, vacancy_by_years.values(), width, label="Количество вакансий")
        ax.bar(x_list2, vacancy_by_year_profession.values(), width,
               label=f'Количество вакансий\n{name_of_profession.lower()}')

        ax.set_xticks(x_nums, vacancy_by_years.keys(), rotation="vertical")
        ax.legend(fontsize=8, borderpad=0)
        ax.grid(True, axis="y")

    def draw_city_salary_graph(self, fig, salary_by_city):
        """Создаёт диаграмму по уровню зарплат по городам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            salary_by_city(dict<string, float>): Статистика уровня зарплат по городам
        """
        width = 0.7
        x_nums = np.arange(len(salary_by_city.keys()))
        x_list = x_nums

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        ax.barh(x_list, list(reversed(salary_by_city.values())), width)

        cities = list(reversed(list(map(lambda item: self.format_city(item), salary_by_city.keys()))))
        ax.set_yticks(x_nums, cities)
        ax.tick_params(axis="y", labelsize=6)
        ax.tick_params(axis="x", labelsize=8)
        ax.grid(True, axis="x")

    def draw_vacancy_rate(self, fig, vacancy_rate):
        """Создаёт диаграмму по доле вакансий по городам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            vacancy_rate(dict<string, float>): Статистика количества вакансий по городам
        """
        vacancy_rate["Другие"] = 1 - sum(vacancy_rate.values())
        vacancy_rate = dict(sorted(vacancy_rate.items(), key=lambda item: item[1]))
        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        ax.pie(vacancy_rate.values(), labels=vacancy_rate.keys(), textprops={'fontsize': 6})
        ax.axis("equal")

    def generate_exel(self, year_statistics, city_statistics, name_of_profession):
        """Создаёт таблицу Exel на основе данных из статистики
        Args:  
            vacancy_by_years(dict<int, int>): Статистика количества вакансий по годам
            vacancy_by_year_profession(dict<int, int>):  Статистика количества вакансий профессии по годам
            name_of_profession(str): Название профессии
        """
        self.create_year_sheet(year_statistics, name_of_profession)
        self.create_city_sheet(city_statistics)
        self.wb.save('report.xlsx')

    def create_year_sheet(self, year_statistic, name_of_profession):
        """Создаёт страницу в таблице exel, связанные с доле вакансий и зарплат
        Args:  
            year_statistic(list<dict<int, int>>): Статистики  связанные с уровнем зарплат и долей вакансий
            name_of_profession(str): Название профессии
        """
        sheet_year = self.wb.active
        sheet_year.title = 'Статистика по годам'
        heades = ['Год', "Средняя зарплата", f'Средняя зарплата - {name_of_profession}', 'Количество вакансий',
                  f'Количество вакансий - {name_of_profession}']
        thin = Side(border_style="thin", color="000000")
        for i, head in enumerate(heades):
            sheet_year.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)
        for year, value in year_statistic[0].items():
            if (int(year) >= 2007 and int(year) <= 2022):
                sheet_year.append([int(year), int(value), int(year_statistic[1][year]), int(year_statistic[2][year]),
                                   int(year_statistic[3][year])])
        self.make_same_width(sheet_year)
        self.make_border(sheet_year, thin)

    def create_city_sheet(self, city_statistics):
        """Создаёт страницу в таблице exel, связанные с доле вакансий и зарплат
        Args:
            city_statistics(list<dict<str, float>>): Статистики  связанные с уровнем зарплат и долей вакансий
        """
        sheet_city = self.wb.create_sheet("Статистика по городам")
        thin = Side(border_style="thin", color="000000")
        sheet_city.cell(row=1, column=1, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=2, value='Уровень зарплат').font = Font(bold=True)
        sheet_city.cell(row=1, column=4, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=5, value='Доля вакансий').font = Font(bold=True)
        for city, value in city_statistics[0].items():
            sheet_city.append([city, int(value)])
        row = 2
        column = 4
        for city, value in city_statistics[1].items():
            sheet_city.cell(row=row, column=column, value=city)
            sheet_city.cell(row=row, column=column + 1, value=float(value)).number_format = \
                openpyxl.styles.numbers.BUILTIN_FORMATS[10]
            row += 1
        self.make_same_width(sheet_city)
        self.make_border(sheet_city, thin)

    def make_same_width(self, sheet):
        """Создаёт столбцам одинаковую ширину
        Args:
            sheet(sheet): страница в exel
        """
        for column in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 1

    def make_border(self, sheet, sideStyle):
        """Задаёт столбцам границу
        Args:
            sheet(sheet): страница в exel
            sideStyle(sideStyle): стиль границы для ячеек
        """
        for column in sheet.columns:
            for cell in column:
                cell.border = Border(left=sideStyle, right=sideStyle, top=sideStyle, bottom=sideStyle)

    def as_text(self, val):
        """Проверяет, что существует ли значение
        Args:
            val(str):значение
        Returns:
            str: Значение
        """
        if val is None:
            return ""
        return str(val)

    def filter_statistics(self, statistics, start_year, end_year):
        """Фильтрует знаечения по годам
        Args:
            statistics<list<int, int>>: Статистика
            start_year(int): Начальный год
            end_year(int): Конечный год
        """
        for statisctic in statistics:
            dict(filter(lambda item: int(item[1]) >= start_year and int(item[1]) <= end_year, statisctic.items()))

    def format_city(self, city):
        """Приводит названте города к стандратному виду
        Args:
            city(str): Название города
        Returns:
            str: Форматированное название города
        """
        if ('-' in city):
            return city.replace('-', '\n')
        if (' - ' in city):
            return city.replace(' - ', '\n')
        if (' ' in city):
            return city.replace(' ', '\n')
        return city


data_type = input('Выберите метод работы: ')
dataset = DataSet()

if (data_type == 'Вакансии'):
    dataset.print_table()

elif (data_type == 'Статистика'):
    statistical_data_processor = StatisticalDataProcessor(dataset)
    report = Report()
    report.generate_image(statistical_data_processor.get_final_year_statistics(),
                          statistical_data_processor.get_final_city_statistics(),
                          statistical_data_processor.name_of_profession)
