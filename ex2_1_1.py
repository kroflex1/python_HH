import csv
import re
import sys
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import openpyxl.styles.numbers
from datetime import datetime


class Salary:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = float(salary_from) * self.currency_to_rub[salary_currency]
        self.salary_to = float(salary_to) * self.currency_to_rub[salary_currency]
        self.salary_currency = salary_currency
        self.average = int((self.salary_from + self.salary_to) / 2)


class Vacancy:
    def __init__(self, vacancy_information):
        self.name = vacancy_information['name']
        self.salary = Salary(vacancy_information['salary_from'], vacancy_information['salary_to'],
                             vacancy_information['salary_currency'])
        self.area_name = vacancy_information['area_name']
        self.published_at = datetime.strptime(vacancy_information['published_at'], '%Y-%m-%dT%H:%M:%S%z')
        self.salary_currency = vacancy_information['salary_currency']


class StatisticalDataProcessor:
    def __init__(self, dataset):
        self.dataset = dataset
        self.name_of_profession = dataset.name_of_profession

        self.average_salary = self.get_salary_statistics()
        self.average_salary_profession = self.get_salary_statistics(self.name_of_profession)
        self.number_of_vacancies = self.get_number_of_vacancies_statistics()
        self.number_of_vacancies_profession = self.get_number_of_vacancies_statistics(self.name_of_profession)
        self.salary_level = self.get_salary_level_city_statistics()
        self.vacancy_rate = self.get_number_of_vacancies_city_statistics()

    def print_statistic(self):
        print('Динамика уровня зарплат по годам: {' + ', '.join(self.average_salary) + '}')
        print('Динамика количества вакансий по годам: {' + ', '.join(self.number_of_vacancies) + '}')
        print('Динамика уровня зарплат по годам для выбранной профессии: {' + ', '.join(
            self.average_salary_profession) + '}')
        print('Динамика количества вакансий по годам для выбранной профессии: {' + ', '.join(
            self.number_of_vacancies_profession) + '}')
        print('Уровень зарплат по городам (в порядке убывания): {' + ', '.join(self.salary_level) + '}')
        print('Доля вакансий по городам (в порядке убывания): {' + ', '.join(self.vacancy_rate) + '}')

    def get_final_year_statistics(self):
        return [self.convert_year_statistic_to_dictionary(self.average_salary),
                self.convert_year_statistic_to_dictionary(self.average_salary_profession),
                self.convert_year_statistic_to_dictionary(self.number_of_vacancies),
                self.convert_year_statistic_to_dictionary(self.number_of_vacancies_profession)]

    def get_final_city_statistics(self):
        return [self.convert_city_statistic_to_dictionary(self.salary_level),
                self.convert_city_statistic_to_dictionary(self.vacancy_rate)]

    def get_salary_statistics(self, name_of_profession=""):
        salary_statistics = []
        current_year = self.dataset.vacancies_objects[0].published_at.year
        average_salary = 0
        number_of_vacancies = 0
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.published_at.year != current_year and (
                    name_of_profession == "" or name_of_profession in vacancy.name):
                if number_of_vacancies == 0:
                    salary_statistics.append(f'{current_year}: {0}')
                else:
                    salary_statistics.append(f'{current_year}: {math.floor(average_salary / number_of_vacancies)}')
                current_year = vacancy.published_at.year
                average_salary = vacancy.salary.average
                number_of_vacancies = 1
            elif name_of_profession == "" or name_of_profession in vacancy.name:
                average_salary += vacancy.salary.average
                number_of_vacancies += 1
        if number_of_vacancies == 0:
            salary_statistics.append(f'{current_year}: {0}')
        else:
            salary_statistics.append(f'{current_year}: {math.floor(average_salary / number_of_vacancies)}')
        return salary_statistics

    def get_number_of_vacancies_statistics(self, name_of_profession=""):
        number_of_vacancies_statistics = []
        current_year = self.dataset.vacancies_objects[0].published_at.year
        number_of_vacancies = 0
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.published_at.year != current_year and (
                    name_of_profession == "" or name_of_profession in vacancy.name):
                number_of_vacancies_statistics.append(f'{current_year}: {number_of_vacancies}')
                current_year = vacancy.published_at.year
                number_of_vacancies = 1
            elif name_of_profession == "" or name_of_profession in vacancy.name:
                number_of_vacancies += 1
        number_of_vacancies_statistics.append(f'{current_year}: {number_of_vacancies}')
        return number_of_vacancies_statistics

    def get_salary_level_city_statistics(self):
        number_of_vacancies = len(self.dataset.vacancies_objects)
        salary_at_cities = {}
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.area_name not in salary_at_cities:
                salary_at_cities[vacancy.area_name] = (1, vacancy.salary.average)
            else:
                salary_at_cities[vacancy.area_name] = (salary_at_cities[vacancy.area_name][0] + 1,
                                                       salary_at_cities[vacancy.area_name][1] + vacancy.salary.average)
        salary_at_cities = dict(
            sorted(salary_at_cities.items(), key=lambda item: item[1][1] / item[1][0], reverse=True))
        average_salary_at_cities = []
        for city_name, value in salary_at_cities.items():
            if value[0] >= math.floor(number_of_vacancies / 100):
                average_salary_at_cities.append(f"'{city_name}': {math.floor(value[1] / value[0])}")
        return average_salary_at_cities[:10]

    def get_number_of_vacancies_city_statistics(self):
        number_of_all_vacancies = len(self.dataset.vacancies_objects)
        number_of_vacancies_at_cities = {}
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.area_name not in number_of_vacancies_at_cities:
                number_of_vacancies_at_cities[vacancy.area_name] = 1
            else:
                number_of_vacancies_at_cities[vacancy.area_name] += 1
        salary_at_cities = dict(sorted(number_of_vacancies_at_cities.items(), key=lambda item: item[1], reverse=True))
        result = []
        for city_name, number_of_vacancies in salary_at_cities.items():
            if number_of_vacancies >= math.floor(number_of_all_vacancies / 100):
                result.append(f"'{city_name}': {round(number_of_vacancies / number_of_all_vacancies, 4)}")
        return result[:10]

    def convert_year_statistic_to_dictionary(self, statistic):
        return {i.split(': ')[0]: i.split(': ')[1] for i in statistic}

    def convert_city_statistic_to_dictionary(self, statistic):
        return {i.split(': ')[0][1:-1]: i.split(': ')[1] for i in statistic}


class DataSet:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.name_of_profession = input('Введите название профессии: ')
        self.vacancies_objects = self.__read_data_vacancies()

    def __read_data_vacancies(self):
        list_naming, list_data = DataSet.csv_reader(self.file_name)
        data_vacancies = DataSet.csv_filer(list_data, list_naming)
        return [Vacancy(vacancy_information) for vacancy_information in data_vacancies]

    @staticmethod
    def csv_reader(file_name):
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        list_data = [x for x in reader_csv]
        if len(list_data) == 0:
            print("Пустой файл")
            sys.exit()
        return list_data[0], list_data[1:]

    @staticmethod
    def csv_filer(reader, list_naming):
        rows = [x for x in reader if len(x) == len(list_naming) and '' not in x]
        dictionary_rows = [DataSet.convert_row_to_dictionary(list_naming, row) for row in rows]
        return dictionary_rows

    @staticmethod
    def convert_row_to_dictionary(columns_name, row):
        dict = {}
        for i in range(len(columns_name)):
            dict[columns_name[i]] = DataSet.convert_cell_to_standard(row[i])
        return dict

    @staticmethod
    def convert_cell_to_standard(cell):
        words = re.sub(r'<[^>]*>', '', cell).split('\n')
        for i in range(len(words)):
            words[i] = ' '.join(words[i].split())
        return ';'.join(map(str, words))


class Report:
    def __init__(self, ):
        self.wb = Workbook()

    def generate_exel(self, year_statistics, city_statistics, name_of_profession):
        self.create_year_sheet(year_statistics, name_of_profession)
        self.create_city_sheet(city_statistics)
        self.wb.save('report.xlsx')

    def create_year_sheet(self, year_statistic, name_of_profession):
        sheet_year = self.wb.active
        sheet_year.title = 'Статистика по годам'
        heades = ['Год', "Средняя зарплата", f'Средняя зарплата - {name_of_profession}', 'Количество вакансий',
                  f'Количество вакансий - {name_of_profession}']
        thin = Side(border_style="thin", color="000000")
        for i, head in enumerate(heades):
            sheet_year.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)
        for year, value in year_statistic[0].items():
            if (int(year) >= 2007 and int(year) <= 2022):
                sheet_year.append([int(year), int(value), int(year_statistic[1][year]), int(year_statistic[2][year]),
                                   int(year_statistic[3][year])])
        self.make_same_width(sheet_year)
        self.make_border(sheet_year, thin)

    def create_city_sheet(self, city_statistics):
        sheet_city = self.wb.create_sheet("Статистика по городам")
        thin = Side(border_style="thin", color="000000")
        sheet_city.cell(row=1, column=1, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=2, value='Уровень зарплат').font = Font(bold=True)
        sheet_city.cell(row=1, column=4, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=5, value='Доля вакансий').font = Font(bold=True)
        for city, value in city_statistics[0].items():
            sheet_city.append([city, int(value)])
        row = 2
        column = 4
        for city, value in city_statistics[1].items():
            sheet_city.cell(row=row, column=column, value=city)
            sheet_city.cell(row=row, column=column + 1, value=float(value)).number_format = \
                openpyxl.styles.numbers.BUILTIN_FORMATS[10]
            row += 1
        self.make_same_width(sheet_city)
        self.make_border_for_city_sheet(sheet_city, thin)

    def make_same_width(self, sheet):
        for column in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 1

    def make_border(self, sheet, sideStyle):
        for column in sheet.columns:
            for cell in column:
                cell.border = Border(left=sideStyle, right=sideStyle, top=sideStyle, bottom=sideStyle)

    def make_border_for_city_sheet(self, sheet, sideStyle):
        for i, column in enumerate(sheet.columns):
            if i == 2:
                continue
            for cell in column:
                cell.border = Border(left=sideStyle, right=sideStyle, top=sideStyle, bottom=sideStyle)

    def as_text(self, val):
        if val is None:
            return ""
        return str(val)


dataset = DataSet()
statistical_data_processor = StatisticalDataProcessor(dataset)
statistical_data_processor.print_statistic()

report = Report()
report.generate_exel(statistical_data_processor.get_final_year_statistics(),
                     statistical_data_processor.get_final_city_statistics(),
                     statistical_data_processor.name_of_profession)
