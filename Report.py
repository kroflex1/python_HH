from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import openpyxl.styles.numbers
import numpy as np
import matplotlib.pyplot as plt

from jinja2 import Environment, FileSystemLoader
import pdfkit

class Report:
    """Класс для визуального представления информации о вакансиях.

        Attributes:
            wb(Workbook)
    """

    def __init__(self):
        """Инициализирует объекты Report"""
        self.wb = Workbook()

    def generate_pdf(self, year_statistics, city_statistics, name_of_profession):
        self.generate_image(year_statistics, city_statistics, name_of_profession)
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {name_of_profession}', 'Количество вакансий',
                  f'Количество вакансий - {name_of_profession}']
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        city_statistics[1].pop('Другие')
        for key, value in city_statistics[1].items():
            city_statistics[1][key] = ('%.2f' % (value * 100) + '%').replace('.',',')

        env = Environment(loader=FileSystemLoader(''))
        template = env.get_template("pattern.html")
        pdf_template = template.render({'name_of_profession': name_of_profession,
                                        'heads1': heads1,
                                        'heads2': heads2,
                                        'salary_by_years': year_statistics[0],
                                        'vac_salary_by_years': year_statistics[1],
                                        'vacs_by_years': year_statistics[2],
                                        'vac_count_by_years': year_statistics[3],
                                        'city_salary': city_statistics[0],
                                        'vacancy_rate': city_statistics[1]
                                        })

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config, options={"enable-local-file-access": None})

    def generate_image(self, year_statistics, city_statistics, name_of_profession):
        """Создаёт изображение с диаграммами по вакансиям
        Args:
            year_statistics(list<dict<int, int>>): Статистики по зарплатам и доли вакансий
            city_statistics(list<dict<int, float>>): Статистики по городам
            name_of_profession(str): Название профессии
        """
        fig = plt.figure()
        self.filter_statistics(year_statistics, 2007, 2022)
        self.draw_year_salary_graph(fig, year_statistics[0], year_statistics[1], name_of_profession)
        self.draw_year_vacancy_graph(fig, year_statistics[2], year_statistics[3], name_of_profession)
        self.draw_city_salary_graph(fig, city_statistics[0])
        self.draw_vacancy_rate(fig, city_statistics[1])

        plt.tight_layout()
        plt.savefig("graph.png", dpi=200)

    def draw_year_salary_graph(self, fig, salary_by_years, salary_by_year_profession, name_of_profession):
        """Создаёт диаграмму по уровню зарплат по годам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            salary_by_years(dict<int, int>): Статистика уровня зарплат по годам
            salary_by_year_profession(dict<int, int>):  Статистика уровня зарплат профессии по годам
            name_of_profession(str): Название профессии
        """
        width = 0.4
        x_nums = np.arange(len(salary_by_years.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, salary_by_years.values(), width, label="средняя з/п")
        ax.bar(x_list2, salary_by_year_profession.values(), width, label=f'з/п {name_of_profession.lower()}')

        ax.tick_params(axis="both", labelsize=8)
        ax.set_xticks(x_nums, salary_by_years.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

    def draw_year_vacancy_graph(self, fig, vacancy_by_years, vacancy_by_year_profession, name_of_profession):
        """Создаёт диаграмму по количеству вакансий по годам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            vacancy_by_years(dict<int, int>): Статистика количества вакансий по годам
            vacancy_by_year_profession(dict<int, int>):  Статистика количества вакансий профессии по годам
            name_of_profession(str): Название профессии
        """
        width = 0.4
        x_nums = np.arange(len(vacancy_by_years.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(222)
        ax.set_title("Количество ваканский по годам")
        ax.bar(x_list1, vacancy_by_years.values(), width, label="Количество вакансий")
        ax.bar(x_list2, vacancy_by_year_profession.values(), width,
               label=f'Количество вакансий\n{name_of_profession.lower()}')

        ax.set_xticks(x_nums, vacancy_by_years.keys(), rotation="vertical")
        ax.legend(fontsize=8, borderpad=0)
        ax.grid(True, axis="y")

    def draw_city_salary_graph(self, fig, salary_by_city):
        """Создаёт диаграмму по уровню зарплат по городам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            salary_by_city(dict<string, float>): Статистика уровня зарплат по городам
        """
        width = 0.7
        x_nums = np.arange(len(salary_by_city.keys()))
        x_list = x_nums

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        ax.barh(x_list, list(reversed(salary_by_city.values())), width)

        cities = list(reversed(list(map(lambda item: self.format_city(item), salary_by_city.keys()))))
        ax.set_yticks(x_nums, cities)
        ax.tick_params(axis="y", labelsize=6)
        ax.tick_params(axis="x", labelsize=8)
        ax.grid(True, axis="x")

    def draw_vacancy_rate(self, fig, vacancy_rate):
        """Создаёт диаграмму по доле вакансий по городам
        Args:
            fig(figure): Фигура, на которой рисуется диаграмма
            vacancy_rate(dict<string, float>): Статистика количества вакансий по городам
        """
        vacancy_rate["Другие"] = 1 - sum(vacancy_rate.values())
        vacancy_rate = dict(sorted(vacancy_rate.items(), key=lambda item: item[1]))
        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        ax.pie(vacancy_rate.values(), labels=vacancy_rate.keys(), textprops={'fontsize': 6})
        ax.axis("equal")

    def generate_exel(self, year_statistics, city_statistics, name_of_profession):
        """Создаёт таблицу Exel на основе данных из статистики
        Args:
            vacancy_by_years(dict<int, int>): Статистика количества вакансий по годам
            vacancy_by_year_profession(dict<int, int>):  Статистика количества вакансий профессии по годам
            name_of_profession(str): Название профессии
        """
        self.create_year_sheet(year_statistics, name_of_profession)
        self.create_city_sheet(city_statistics)
        self.wb.save('report.xlsx')

    def create_year_sheet(self, year_statistic, name_of_profession):
        """Создаёт страницу в таблице exel, связанные с доле вакансий и зарплат
        Args:
            year_statistic(list<dict<int, int>>): Статистики  связанные с уровнем зарплат и долей вакансий
            name_of_profession(str): Название профессии
        """
        sheet_year = self.wb.active
        sheet_year.title = 'Статистика по годам'
        heades = ['Год', "Средняя зарплата", f'Средняя зарплата - {name_of_profession}', 'Количество вакансий',
                  f'Количество вакансий - {name_of_profession}']
        thin = Side(border_style="thin", color="000000")
        for i, head in enumerate(heades):
            sheet_year.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)
        for year, value in year_statistic[0].items():
            if (int(year) >= 2007 and int(year) <= 2022):
                sheet_year.append([int(year), int(value), int(year_statistic[1][year]), int(year_statistic[2][year]),
                                   int(year_statistic[3][year])])
        self.make_same_width(sheet_year)
        self.make_border(sheet_year, thin)

    def create_city_sheet(self, city_statistics):
        """Создаёт страницу в таблице exel, связанные с доле вакансий и зарплат
        Args:
            city_statistics(list<dict<str, float>>): Статистики  связанные с уровнем зарплат и долей вакансий
        """
        sheet_city = self.wb.create_sheet("Статистика по городам")
        thin = Side(border_style="thin", color="000000")
        sheet_city.cell(row=1, column=1, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=2, value='Уровень зарплат').font = Font(bold=True)
        sheet_city.cell(row=1, column=4, value='Город').font = Font(bold=True)
        sheet_city.cell(row=1, column=5, value='Доля вакансий').font = Font(bold=True)
        for city, value in city_statistics[0].items():
            sheet_city.append([city, int(value)])
        row = 2
        column = 4
        for city, value in city_statistics[1].items():
            sheet_city.cell(row=row, column=column, value=city)
            sheet_city.cell(row=row, column=column + 1, value=float(value)).number_format = \
                openpyxl.styles.numbers.BUILTIN_FORMATS[10]
            row += 1
        self.make_same_width(sheet_city)
        self.make_border(sheet_city, thin)

    def make_same_width(self, sheet):
        """Создаёт столбцам одинаковую ширину
        Args:
            sheet(sheet): страница в exel
        """
        for column in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 1

    def make_border(self, sheet, sideStyle):
        """Задаёт столбцам границу
        Args:
            sheet(sheet): страница в exel
            sideStyle(sideStyle): стиль границы для ячеек
        """
        for column in sheet.columns:
            for cell in column:
                cell.border = Border(left=sideStyle, right=sideStyle, top=sideStyle, bottom=sideStyle)

    def as_text(self, val):
        """Проверяет, что существует ли значение
        Args:
            val(str):значение
        Returns:
            str: Значение
        """
        if val is None:
            return ""
        return str(val)

    def filter_statistics(self, statistics, start_year, end_year):
        """Фильтрует знаечения по годам
        Args:
            statistics<list<int, int>>: Статистика
            start_year(int): Начальный год
            end_year(int): Конечный год
        """
        for statisctic in statistics:
            dict(filter(lambda item: int(item[1]) >= start_year and int(item[1]) <= end_year, statisctic.items()))

    @staticmethod
    def format_city(city):
        """Приводит названте города к стандратному виду
        Args:
            city(str): Название города
        Returns:
            str: Форматированное название города
        """
        if ('-' in city and ' - ' not in city):
            return city.replace('-', '\n')
        if (' - ' in city):
            return city.replace(' - ', '\n')
        if (' ' in city):
            return city.replace(' ', '\n')
        return city